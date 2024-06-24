import   openpyxl
wb = openpyxl.load_workbook('videogamesales.xlsx') #CArgar el libro EXCEL
ws = wb.active #sE DEFINE LA HOJA ACTIVA
ws = wb['vgsales']# SE DEFINE LA HOJA Q USARIAMOS
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns:'+str(ws.max_column))
print('The value in cell A1 is: '+ws['A1'].value) #Para recuperar datos de una celda concreta con Openpyxl, puedes escribir el valor de la celda de la siguiente manera
values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)

lista = []
for i in range (1, ws.max_column+1) :
    values = ws.cell(row=1, column=i).value 
    lista.append(values)
print(values)
data=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)

lista = []
for i in range (2,12) :
    values = ws.cell(row=i, column=2).value 
    lista.append(data)

my_list =[]
for value in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col= 6, values_only=True) :
    my_list.append(value)

print(my_list)
for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    (print ("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6)))

ws['K1'] = 'Sum of Sales'
ws.cell(row=1, column=11, value='Sum of Sales')
wb.save('videogamesales.xlsx')
